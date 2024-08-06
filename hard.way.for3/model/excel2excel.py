# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from datetime import datetime
import requests

# 读取Excel文件
df = pd.read_excel("C:\\Users\\admin\\Desktop\\退订1.xlsx")


def send_post_request(url, data):
    start_time = datetime.now()
    response = requests.post(url, json=data)
    end_time = datetime.now()
    print(f"{url}  Total time: {(end_time - start_time)} ms")
    return response.json()


# 创建一个新的excel文件
workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
for i in range(len(df)):
    half1 = df.iloc[i]['id']
    txt = df.iloc[i]['txt']
    data = {"stream": False, "model": "Qwen2-7b", "audio_text": txt}
    response = send_post_request("http://192.168.2.9:8083/v1/chat/quality_inspection", data)
    response2 = send_post_request("http://192.168.2.58:8083/v1/chat/newqc", data)

    sheet.cell(row, 1, half1)
    sheet.cell(row, 2, txt)
    sheet.cell(row, 3, response['data']['message'])
    sheet.cell(row, 4, response2['data']['message'])
    row += 1

# 保存excel文件
workbook.save("C:\\Users\\admin\\Desktop\\退订2.xlsx")
