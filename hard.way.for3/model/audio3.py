# -*- coding: utf-8 -*-
import os

import requests
import pandas as pd
from openpyxl import Workbook


# 下载音频文件
def download_audio(url, filename):
    # 检查文件是否已存在
    if os.path.exists(filename):
        print(f"文件 '{filename}' 已存在，跳过下载。")
        return

    # 如果文件不存在，则开始下载
    try:
        response = requests.get(url)
        response.raise_for_status()  # 检查请求是否成功

        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f"文件 '{filename}' 下载完成。")
    except requests.exceptions.RequestException as e:
        print(f"下载失败: {e}")


# 创建一个新的工作簿
wb = Workbook()
sheet = wb.active
df = pd.read_excel("D:\\data\\convert16\\过程声纹_data.xlsx")

uids = [14571,
        1577,
        22889,
        23851,
        22765,
        24263,
        1905,
        1903,
        25446,
        26098,
        2329,
        2221,
        25213,
        25338,
        11737,
        2213,
        1927,
        24696,
        2591,
        23247,
        1379,
        2467,
        26870,
        1377,
        12239,
        23815,
        20781,
        8467,
        25314,
        22167,
        25313,
        2237,
        1711,
        24793,
        24792,
        27387,
        8483,
        24205,
        14529,
        2255,
        25090,
        25093,
        12107,
        26839,
        2151,
        16915,
        25066,
        25068,
        2145,
        16359,
        12677,
        2163,
        1985,
        25074,
        10707,
        14743,
        1507,
        1627,
        1869,
        25071,
        25070,
        26939,
        26938,
        22577,
        22457,
        25284,
        25164,
        1999,
        19051,
        26942,
        20964,
        26263,
        1641,
        11455,
        1647,
        17559,
        20497,
        25146,
        25260,
        15905,
        1893,
        16435,
        15467,
        1537,
        22449,
        25279,
        26369,
        11315,
        22535,
        1551,
        21331,
        23871,
        25135,
        25372
        ]

count = 1
for i in range(len(df)):
    name = df.iloc[i]['audio_test']
    dist = df.iloc[i]['dist']
    uid = df.iloc[i]['uid']

    if int(uid) in uids and (0.5 < float(dist) <= 0.6):
        sheet.cell(count, 1, uid)
        sheet.cell(count, 3, dist)
        sheet.cell(count, 4, name)

        download_audio(f"http://192.168.2.4:12000/audio_test/{name}",
                       f"D:\\data\\convert16\\audio_test\\{name}")
        download_audio(f"http://192.168.2.4:12000/audio_testdb/{uid}.wav",
                       f"D:\\data\\convert16\\audio_testdb\\{uid}.wav")

        attachment_path = 'audio_test\\' + name
        sheet[f'B{count}'] = f'=HYPERLINK("{attachment_path}", "audio_test {name}")'
        attachment_path = 'audio_testdb\\' + str(uid) + '.wav'
        sheet[f'E{count}'] = f'=HYPERLINK("{attachment_path}", "audio_testdb {uid}.wav")'
        count += 1
        print(count)


# 保存Excel文件
wb.save("D:\\data\\convert16\\audio.xlsx")
