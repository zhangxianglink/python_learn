# -*- coding: utf-8 -*-
import json
import openpyxl

# 读取json文件
with open('C:\\Users\\admin\\Desktop\\过程声纹.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

# 创建一个新的excel文件
workbook = openpyxl.Workbook()
sheet = workbook.active

# 遍历键值对并写入excel文件
row = 1
for item in data:
    audio_test = item['audio_test']
    user_id = item['user_id']
    result = item['dist']

    data = json.loads(result)

    if 'dist' in data:
        sheet.cell(row, 1, audio_test)
        sheet.cell(row, 2, data['dist'])
        sheet.cell(row, 3, user_id)
        sheet.cell(row, 4, data['info'])
        row += 1


# 保存excel文件
workbook.save(f"C:\\Users\\admin\\Desktop\\过程声纹{row}.xlsx")

