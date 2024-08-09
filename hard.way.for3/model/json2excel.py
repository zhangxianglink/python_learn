# -*- coding: utf-8 -*-
import json
import openpyxl

# 读取json文件
with open("C:\\Users\\admin\\Desktop\\标数\\0807\\偏差125.json", 'r', encoding='utf-8') as file:
    data = json.load(file)

# 创建一个新的excel文件
workbook = openpyxl.Workbook()
sheet = workbook.active

# 遍历键值对并写入excel文件
row = 1
for item in data:
    output = item['output']
    input = item['input']
    instruction = item['instruction']


    sheet.cell(row, 1, input)
    sheet.cell(row, 2, instruction)
    sheet.cell(row, 3, json.dumps(output, ensure_ascii=False, indent=4))


    row += 1

# 保存excel文件
workbook.save("C:\\Users\\admin\\Desktop\\标数\\0807\\偏差125.xlsx")

